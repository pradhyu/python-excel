"""Excel file loading functionality with support for multiple formats."""

import os
import pandas as pd
from pathlib import Path
from typing import Dict, List, Union, Optional
from datetime import datetime

from .exceptions import FileLoadError, ExcelProcessorError
from .models import ExcelFile


class ExcelLoader:
    """Handles loading Excel files and converting them to DataFrames."""
    
    def __init__(self):
        self.supported_extensions = {'.xlsx', '.xls', '.xlsm', '.xlsb'}
    
    def validate_file(self, file_path: Union[str, Path]) -> bool:
        """Validate if the file exists and is a supported Excel format."""
        file_path = Path(file_path)
        
        if not file_path.exists():
            return False
        
        if not file_path.is_file():
            return False
        
        if file_path.suffix.lower() not in self.supported_extensions:
            return False
        
        return True
    
    def get_sheet_names(self, file_path: Union[str, Path]) -> List[str]:
        """Get list of sheet names from an Excel file."""
        file_path = Path(file_path)
        
        if not self.validate_file(file_path):
            raise FileLoadError(
                str(file_path),
                "File validation failed",
                "File does not exist or is not a supported Excel format"
            )
        
        try:
            if file_path.suffix.lower() == '.xls':
                # Use xlrd for older .xls files
                import xlrd
                workbook = xlrd.open_workbook(str(file_path))
                return workbook.sheet_names()
            else:
                # Use openpyxl for newer formats
                from openpyxl import load_workbook
                workbook = load_workbook(str(file_path), read_only=True)
                return workbook.sheetnames
        
        except Exception as e:
            error_msg = str(e)
            if "zip file" in error_msg.lower() or "not a valid" in error_msg.lower():
                raise FileLoadError(
                    str(file_path),
                    "Failed to load Excel file",
                    f"File appears to be corrupted or not a valid Excel file: {error_msg}"
                )
            raise FileLoadError(
                str(file_path),
                "Failed to read sheet names",
                str(e)
            )
    
    def load_sheet(self, file_path: Union[str, Path], sheet_identifier: Union[str, int]) -> pd.DataFrame:
        """Load a specific sheet from an Excel file."""
        file_path = Path(file_path)
        
        if not self.validate_file(file_path):
            raise FileLoadError(
                str(file_path),
                "File validation failed",
                "File does not exist or is not a supported Excel format"
            )
        
        try:
            # Determine the engine based on file extension
            engine = 'xlrd' if file_path.suffix.lower() == '.xls' else 'openpyxl'
            
            # Load the specific sheet
            df = pd.read_excel(
                str(file_path),
                sheet_name=sheet_identifier,
                engine=engine,
                na_values=['', 'N/A', 'NA', 'NULL', 'null', '#N/A', '#NULL!'],
                keep_default_na=True
            )
            
            # Clean up column names (remove leading/trailing whitespace)
            df.columns = df.columns.astype(str).str.strip()
            
            # Infer and optimize data types
            df = self._optimize_dtypes(df)
            
            return df
        
        except Exception as e:
            if isinstance(e, FileLoadError):
                raise
            
            # Handle specific pandas/Excel errors
            error_msg = str(e)
            if "No sheet named" in error_msg or "Worksheet named" in error_msg:
                available_sheets = self.get_sheet_names(file_path)
                raise FileLoadError(
                    str(file_path),
                    f"Sheet '{sheet_identifier}' not found",
                    f"Available sheets: {', '.join(available_sheets)}"
                )
            elif "Worksheet index" in error_msg:
                sheet_count = len(self.get_sheet_names(file_path))
                raise FileLoadError(
                    str(file_path),
                    f"Sheet index {sheet_identifier} out of range",
                    f"File has {sheet_count} sheets (indices 0-{sheet_count-1})"
                )
            else:
                raise FileLoadError(
                    str(file_path),
                    "Failed to load sheet",
                    error_msg
                )
    
    def load_file(self, file_path: Union[str, Path]) -> Dict[str, pd.DataFrame]:
        """Load all sheets from an Excel file."""
        file_path = Path(file_path)
        
        if not self.validate_file(file_path):
            raise FileLoadError(
                str(file_path),
                "File validation failed",
                "File does not exist or is not a supported Excel format"
            )
        
        try:
            sheet_names = self.get_sheet_names(file_path)
            sheets = {}
            
            for sheet_name in sheet_names:
                try:
                    sheets[sheet_name] = self.load_sheet(file_path, sheet_name)
                except FileLoadError as e:
                    # Log the error but continue with other sheets
                    print(f"Warning: Failed to load sheet '{sheet_name}': {e.message}")
                    continue
            
            if not sheets:
                raise FileLoadError(
                    str(file_path),
                    "No sheets could be loaded",
                    "All sheets failed to load or file contains no readable sheets"
                )
            
            return sheets
        
        except Exception as e:
            if isinstance(e, FileLoadError):
                raise
            
            # Check if it's a corrupted file error
            error_msg = str(e)
            if "zip file" in error_msg.lower() or "not a valid" in error_msg.lower():
                raise FileLoadError(
                    str(file_path),
                    "Failed to load Excel file",
                    f"File appears to be corrupted or not a valid Excel file: {error_msg}"
                )
            
            raise FileLoadError(
                str(file_path),
                "Failed to load Excel file",
                str(e)
            )
    
    def create_excel_file_model(self, file_path: Union[str, Path], db_directory: Union[str, Path]) -> ExcelFile:
        """Create an ExcelFile model from a file path."""
        file_path = Path(file_path)
        db_directory = Path(db_directory)
        
        # Load all sheets
        sheets = self.load_file(file_path)
        
        # Calculate memory usage
        memory_usage = sum(df.memory_usage(deep=True).sum() for df in sheets.values()) / (1024 * 1024)  # MB
        
        # Get file modification time
        last_modified = datetime.fromtimestamp(file_path.stat().st_mtime)
        
        # Create relative file name
        try:
            relative_path = file_path.relative_to(db_directory)
            file_name = str(relative_path)
        except ValueError:
            # File is not within db_directory
            file_name = file_path.name
        
        return ExcelFile(
            file_name=file_name,
            file_path=str(file_path),
            sheets=sheets,
            last_modified=last_modified,
            memory_usage=memory_usage
        )
    
    def _optimize_dtypes(self, df: pd.DataFrame) -> pd.DataFrame:
        """Optimize DataFrame data types for memory efficiency."""
        # Make a copy to avoid modifying the original
        df = df.copy()
        
        for column in df.columns:
            col_type = df[column].dtype
            
            # Skip if already optimized or contains mixed types
            if col_type == 'object':
                # Try to convert to numeric if possible
                numeric_series = pd.to_numeric(df[column], errors='coerce')
                if not numeric_series.isna().all():
                    # If most values can be converted to numeric, use numeric type
                    non_null_count = df[column].notna().sum()
                    numeric_count = numeric_series.notna().sum()
                    if numeric_count / non_null_count > 0.8:  # 80% threshold
                        df[column] = numeric_series
                        continue
                
                # Try to convert to datetime
                try:
                    datetime_series = pd.to_datetime(df[column], errors='coerce')
                    if not datetime_series.isna().all():
                        non_null_count = df[column].notna().sum()
                        datetime_count = datetime_series.notna().sum()
                        if datetime_count / non_null_count > 0.8:  # 80% threshold
                            df[column] = datetime_series
                            continue
                except:
                    pass
                
                # Try to convert to category if low cardinality
                unique_count = df[column].nunique()
                total_count = len(df[column])
                if unique_count < total_count * 0.5:  # Less than 50% unique values
                    df[column] = df[column].astype('category')
            
            elif col_type in ['int64', 'float64']:
                # Downcast numeric types
                if col_type == 'int64':
                    df[column] = pd.to_numeric(df[column], downcast='integer')
                elif col_type == 'float64':
                    df[column] = pd.to_numeric(df[column], downcast='float')
        
        return df
    
    def get_file_info(self, file_path: Union[str, Path]) -> Dict[str, any]:
        """Get basic information about an Excel file without loading all data."""
        file_path = Path(file_path)
        
        if not self.validate_file(file_path):
            raise FileLoadError(
                str(file_path),
                "File validation failed",
                "File does not exist or is not a supported Excel format"
            )
        
        try:
            sheet_names = self.get_sheet_names(file_path)
            file_size = file_path.stat().st_size / (1024 * 1024)  # MB
            last_modified = datetime.fromtimestamp(file_path.stat().st_mtime)
            
            return {
                'file_path': str(file_path),
                'file_name': file_path.name,
                'file_size_mb': round(file_size, 2),
                'sheet_names': sheet_names,
                'sheet_count': len(sheet_names),
                'last_modified': last_modified,
                'extension': file_path.suffix.lower()
            }
        
        except Exception as e:
            raise FileLoadError(
                str(file_path),
                "Failed to get file information",
                str(e)
            )